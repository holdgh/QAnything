from qanything_kernel.utils.general_utils import get_time, get_table_infos, num_tokens_embed, get_all_subpages, \
    html_to_markdown, clear_string, get_time_async
from typing import List, Optional
from qanything_kernel.configs.model_config import UPLOAD_ROOT_PATH, LOCAL_OCR_SERVICE_URL, IMAGES_ROOT_PATH, \
    DEFAULT_CHILD_CHUNK_SIZE, LOCAL_PDF_PARSER_SERVICE_URL
from langchain.docstore.document import Document
from qanything_kernel.utils.loader.my_recursive_url_loader import MyRecursiveUrlLoader
from qanything_kernel.utils.custom_log import insert_logger
from langchain_community.document_loaders import UnstructuredFileLoader, TextLoader
from langchain_community.document_loaders import UnstructuredWordDocumentLoader
from langchain_community.document_loaders import UnstructuredEmailLoader
from langchain_community.document_loaders import UnstructuredPowerPointLoader
from qanything_kernel.utils.loader import UnstructuredPaddlePDFLoader
from langchain.text_splitter import RecursiveCharacterTextSplitter
from qanything_kernel.utils.loader.csv_loader import CSVLoader
from qanything_kernel.utils.loader.markdown_parser import convert_markdown_to_langchaindoc
import asyncio
import aiohttp
import docx2txt
import base64
import pandas as pd
import os
import json
import requests
import threading
import re
import newspaper
import uuid
import traceback
import openpyxl
import shutil
import time


def get_ocr_result_sync(image_data):
    try:
        response = requests.post(f"http://{LOCAL_OCR_SERVICE_URL}/ocr", data=image_data, timeout=120)
        response.raise_for_status()  # 如果请求返回了错误状态码，将会抛出异常
        ocr_res = response.text
        ocr_res = json.loads(ocr_res)
        return ocr_res['result']
    except Exception as e:
        insert_logger.warning(f"ocr error: {traceback.format_exc()}")
        return None

def get_pdf_result_sync(file_path):
    try:
        data = {
            'filename': file_path,
            'save_dir': os.path.dirname(file_path)
        }
        headers = {"content-type": "application/json"}
        response = requests.post(f"http://{LOCAL_PDF_PARSER_SERVICE_URL}/pdfparser", json=data, headers=headers,
                                 timeout=240)
        response.raise_for_status()  # 如果请求返回了错误状态码，将会抛出异常
        response_json = response.json()
        markdown_file = response_json.get('markdown_file')
        return markdown_file
    except Exception as e:
        insert_logger.warning(f"pdf parser error: {traceback.format_exc()}")
        return None


class LocalFileForInsert:
    def __init__(self, user_id, kb_id, file_id, file_location, file_name, file_url, chunk_size, mysql_client):
        # 文件切片尺寸
        self.chunk_size = chunk_size
        # 递归切片对象
        self.markdown_text_splitter = RecursiveCharacterTextSplitter(chunk_size=chunk_size, chunk_overlap=0,
                                                                     length_function=num_tokens_embed)
        # 用户id
        self.user_id = user_id
        # 知识库id
        self.kb_id = kb_id
        # 文件id
        self.file_id = file_id
        # 初始化文件切片得到的文档列表
        self.docs: List[Document] = []
        # 初始化词嵌入列表【什么作用？】
        self.embs = []
        # 文件名
        self.file_name = file_name
        # 文件位置
        self.file_location = file_location
        # 文件路由，用以收集链接地址
        self.file_url = ""
        # 问答字典【什么作用？收集问答集信息】
        self.faq_dict = {}
        # 文件路径
        self.file_path = ""
        # MySQL客户端
        self.mysql_client = mysql_client
        if self.file_location == 'FAQ':
            # 当文件位置是FAQ时，表示文件为问答集
            # 获取问答信息
            faq_info = self.mysql_client.get_faq(self.file_id)
            # 从问答信息中提取字段
            user_id, kb_id, question, answer, nos_keys = faq_info
            # 设置问答字段
            self.faq_dict = {'question': question, 'answer': answer, 'nos_keys': nos_keys}
        elif self.file_location == 'URL':
            # 当文件位置是URL，表示链接地址
            # 设置文件路由
            self.file_url = file_url
            # 获取上传地址
            upload_path = os.path.join(UPLOAD_ROOT_PATH, user_id)
            # 获取文件路径
            file_dir = os.path.join(upload_path, self.kb_id, self.file_id)
            # 创建文件路径
            os.makedirs(file_dir, exist_ok=True)
            # 获取文件绝对路径
            self.file_path = os.path.join(file_dir, self.file_name)
        else:
            # 当文件位置非上述两种情况时，直接赋值文件位置为文件路径
            self.file_path = self.file_location
        self.event = threading.Event()

    @staticmethod
    @get_time
    def image_ocr_txt(filepath, dir_path="tmp_files"):
        full_dir_path = os.path.join(os.path.dirname(filepath), dir_path)
        if not os.path.exists(full_dir_path):
            os.makedirs(full_dir_path)
        filename = os.path.split(filepath)[-1]

        # 读取图片
        img_np = open(filepath, 'rb').read()

        img_data = {
            "img64": base64.b64encode(img_np).decode("utf-8"),
        }

        result = get_ocr_result_sync(img_data)

        ocr_result = [line for line in result if line]
        ocr_result = '\n'.join(ocr_result)

        insert_logger.info(f'ocr_res[:100]: {ocr_result[:100]}')

        # 写入结果到文本文件
        txt_file_path = os.path.join(full_dir_path, "%s.txt" % (filename))
        with open(txt_file_path, 'w', encoding='utf-8') as fout:
            fout.write(ocr_result)

        return txt_file_path

    def table_process(self, doc):
        table_infos = get_table_infos(doc.page_content)
        title_lst = doc.metadata['title_lst']
        new_docs = []
        if table_infos is not None:
            tmp_content = '\n'.join(title_lst) + '\n' + doc.page_content
            if num_tokens_embed(tmp_content) <= self.chunk_size:
                doc.page_content = tmp_content
                return [doc]
            head_line = table_infos['head_line']
            end_line = table_infos['end_line']
            table_head = table_infos['head']

            # 处理表格前的内容
            if head_line != 0:
                tmp_doc = Document(
                    page_content='\n'.join(title_lst) + '\n' + '\n'.join(table_infos['lines'][:head_line]),
                    metadata=doc.metadata)
                new_docs.append(tmp_doc)

            # 处理表格内容
            table_content = '\n'.join(title_lst) + '\n' + table_head
            for line in table_infos['lines'][head_line + 2:end_line + 1]:
                if num_tokens_embed(table_content + '\n' + line) > self.chunk_size:
                    # 如果添加新行会超出chunk_size，先保存当前内容
                    tmp_doc = Document(page_content=table_content, metadata=doc.metadata)
                    new_docs.append(tmp_doc)
                    # 重新开始一个新的chunk，包含标题和表头
                    table_content = '\n'.join(title_lst) + '\n' + table_head + '\n' + line
                else:
                    if line == table_head.split('\n')[0]:
                        table_content += '\n\n' + line
                        # print('match table_head:', table_content)
                    else:
                        table_content += '\n' + line

            # 保存最后一个chunk
            if table_content != '\n'.join(title_lst) + '\n' + table_head:
                tmp_doc = Document(page_content=table_content, metadata=doc.metadata)
                new_docs.append(tmp_doc)

            # 处理表格后的内容
            if end_line != len(table_infos['lines']) - 1:
                tmp_doc = Document(
                    page_content='\n'.join(title_lst) + '\n' + '\n'.join(table_infos['lines'][end_line:]),
                    metadata=doc.metadata)
                new_docs.append(tmp_doc)

            insert_logger.info(f"TABLE SLICES: {new_docs[:2]}")
        else:
            return None
        return new_docs

    @staticmethod
    def get_page_id(doc, pre_page_id):
        # 查找 page_id 标志行
        lines = doc.page_content.split('\n')
        for line in lines:
            if re.match(r'^#+ 当前页数:\d+$', line):
                try:
                    page_id = int(line.split(':')[-1])
                    return page_id
                except ValueError:
                    continue
        return pre_page_id

    def markdown_process(self, docs: List[Document]):
        new_docs = []
        for doc in docs:
            title_lst = doc.metadata['title_lst']
            # 删除所有仅有多个#的title
            title_lst = [t for t in title_lst if t.replace('#', '') != '']
            has_table = doc.metadata['has_table']
            if has_table:
                table_doc_id = str(uuid.uuid4())
                self.mysql_client.add_document(table_doc_id, doc.to_json())
                doc.metadata['table_doc_id'] = table_doc_id
                table_docs = self.table_process(doc)
                if table_docs:
                    new_docs.extend(table_docs)
                    continue
            slices = self.markdown_text_splitter.split_documents([doc])
            # insert_logger.info(f"markdown_text_splitter: {len(slices)}")
            if len(slices) == 1:
                slices[0].page_content = '\n\n'.join(title_lst) + '\n\n' + slices[0].page_content
            else:
                for idx, slice in enumerate(slices):
                    slice.page_content = '\n\n'.join(
                        title_lst) + f'\n\n###### 第{idx + 1}段内容如下：\n' + slice.page_content
            new_docs.extend(slices)
        return new_docs

    @get_time_async
    async def url_to_documents_async(self, file_path, file_name, file_url, dir_path="tmp_files", max_retries=3):
        full_dir_path = os.path.join(os.path.dirname(file_path), dir_path)
        if not os.path.exists(full_dir_path):
            os.makedirs(full_dir_path)

        for attempt in range(max_retries):
            try:
                headers = {
                    "Accept": "application/json",
                    "X-Return-Format": "markdown",
                    "X-Timeout": "15",
                }
                async with aiohttp.ClientSession() as session:
                    async with session.get(f"https://r.jina.ai/{file_url}", headers=headers, timeout=30) as response:
                        jina_response = await response.json()
                        if jina_response['code'] == 200:
                            title = jina_response['data'].get('title', '')
                            markdown_str = jina_response['data'].get('content', '')
                            markdown_str = html_to_markdown(markdown_str)
                            md_file_path = os.path.join(full_dir_path, "%s.md" % (file_name))
                            with open(md_file_path, 'w', encoding='utf-8') as fout:
                                fout.write(markdown_str)
                            docs = convert_markdown_to_langchaindoc(md_file_path)
                            if title:
                                for doc in docs:
                                    doc.metadata['title'] = title
                            docs = self.markdown_process(docs)
                            return docs
                        else:
                            insert_logger.warning(f"jina get url warning: {file_url}, {jina_response}")
            except Exception as e:
                insert_logger.warning(f"jina get url error: {file_url}, {traceback.format_exc()}")

            if attempt < max_retries - 1:  # 如果不是最后一次尝试，等待30秒后重试
                await asyncio.sleep(30)

        return None

    @get_time
    def url_to_documents(self, file_path, file_name, file_url, dir_path="tmp_files", max_retries=3):
        """
        由链接地址，借助爬虫，生成文档对象列表
        """
        full_dir_path = os.path.join(os.path.dirname(file_path), dir_path)
        if not os.path.exists(full_dir_path):
            os.makedirs(full_dir_path)

        for attempt in range(max_retries):
            try:
                headers = {
                    "Accept": "application/json",
                    "X-Return-Format": "markdown",
                    "X-Timeout": "15",
                }
                response = requests.get(f"https://r.jina.ai/{file_url}", headers=headers, timeout=30)
                jina_response = response.json()
                if jina_response['code'] == 200:
                    title = jina_response['data'].get('title', '')
                    markdown_str = jina_response['data'].get('content', '')
                    markdown_str = html_to_markdown(markdown_str)
                    md_file_path = os.path.join(full_dir_path, "%s.md" % (file_name))
                    with open(md_file_path, 'w', encoding='utf-8') as fout:
                        fout.write(markdown_str)
                    docs = convert_markdown_to_langchaindoc(md_file_path)
                    if title:
                        for doc in docs:
                            doc.metadata['title'] = title
                    docs = self.markdown_process(docs)
                    return docs
                else:
                    insert_logger.warning(f"jina get url warning: {file_url}, {jina_response}")
            except Exception as e:
                insert_logger.warning(f"jina get url error: {file_url}, {traceback.format_exc()}")

            if attempt < max_retries - 1:  # 如果不是最后一次尝试，等待30秒后重试
                time.sleep(30)

        return None

    @staticmethod
    def excel_to_markdown(file_path, markdown_path):
        def clean_cell_content(cell):
            if cell is None:
                return ''
            # 将单元格内容转换为字符串，并替换换行符为空格
            return re.sub(r'\s+', ' ', str(cell)).strip()
        basename = os.path.splitext(os.path.basename(file_path))[0]
        markdown_file = os.path.join(markdown_path, f"{basename}.md")

        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        with open(markdown_file, 'w', encoding='utf-8') as md_file:
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                md_file.write(f"# {sheet_name}\n\n")

                # 获取非空行和列
                rows = [[clean_cell_content(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]
                non_empty_rows = [row for row in rows if any(cell != '' for cell in row)]

                if not non_empty_rows:
                    continue  # 跳过空表格

                max_cols = max(len(row) for row in non_empty_rows)

                # 处理每一行
                for row_index, row in enumerate(non_empty_rows):
                    # 补齐空单元格
                    padded_row = row + [''] * (max_cols - len(row))

                    # 转换为Markdown表格行，使用竖线作为分隔符
                    markdown_row = '| ' + ' | '.join(padded_row) + ' |'
                    md_file.write(markdown_row + '\n')

                    # 在第一行后添加分隔符
                    if row_index == 0:
                        separator = '|' + '|'.join(['---' for _ in range(max_cols)]) + '|'
                        md_file.write(separator + '\n')

                md_file.write('\n\n')  # 在每个表格后添加空行

        insert_logger.info(f"转换完成。Markdown 文件已保存为 {markdown_file}")
        return markdown_file

    @staticmethod
    def load_text(file_path):
        encodings = ['utf-8', 'iso-8859-1', 'windows-1252']

        for encoding in encodings:
            try:
                loader = TextLoader(file_path, encoding=encoding)
                docs = loader.load()
                insert_logger.info(f"TextLoader {encoding} success: {file_path}")
                return docs
            except Exception:
                insert_logger.warning(f"TextLoader {encoding} error: {file_path}, {traceback.format_exc()}")

        insert_logger.error(f"Failed to load file with all attempted encodings: {file_path}")
        return []

    @staticmethod
    def copy_images(image_root_path, output_dir):
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        # 获取当前目录下所有jpg文件
        images = [f for f in os.listdir(image_root_path) if f.endswith('.jpg')]
        # 复制到指定目录
        for image in images:
            single_image_path = os.path.join(image_root_path, image)
            insert_logger.info(f"copy image: {single_image_path} -> {output_dir}")
            shutil.copy(single_image_path, output_dir)

    @get_time
    def split_file_to_docs(self):
        """
        将单个文件切分成文档列表
        """
        # 打印开始切片日志
        insert_logger.info(f"start split file to docs, file_path: {self.file_name}")
        if self.faq_dict:
            # 问答字典非空，表示问答集模式
            # 直接利用文档字典构造文档对象列表
            docs = [Document(page_content=self.faq_dict['question'], metadata={"faq_dict": self.faq_dict})]
        elif self.file_url:
            # 链接地址非空，表示链接，借助爬虫生成文档
            insert_logger.info("load url: {}".format(self.file_url))
            docs = self.url_to_documents(self.file_path, self.file_name, self.file_url)
            if docs is None:
                # 如果爬虫所得文档列表为空
                try:
                    # 使用newspaper工具，传入链接地址，获取解析文章
                    article = newspaper.article(self.file_url, timeout=120)
                    # 利用文章构造文档对象列表
                    docs = [
                        Document(page_content=article.text, metadata={"title": article.title, "url": self.file_url})]
                except Exception as e:
                    # 利用newspaper工具异常
                    # 打印异常日志
                    insert_logger.error(f"newspaper get url error: {self.file_url}, {traceback.format_exc()}")
                    # 递归获取url开头的所有链接？
                    loader = MyRecursiveUrlLoader(url=self.file_url)
                    docs = loader.load()
        elif self.file_path.lower().endswith(".md"):
            # markdown文件
            try:
                # 将markdown文件处理为文档列表
                docs = convert_markdown_to_langchaindoc(self.file_path)
                docs = self.markdown_process(docs)
            except Exception as e:
                # markdown文件切片异常
                # 打印异常日志
                insert_logger.error(
                    f"convert_markdown_to_langchaindoc error: {self.file_path}, {traceback.format_exc()}")
                # 利用非结构化文件加载器获取文档列表
                loader = UnstructuredFileLoader(self.file_path, strategy="fast")
                docs = loader.load()
        elif self.file_path.lower().endswith(".txt"):
            # 文本文件，txt文件
            docs = self.load_text(self.file_path)
        elif self.file_path.lower().endswith(".pdf"):
            # pdf文件
            # 将pdf文件转换为markdown文件
            markdown_file = get_pdf_result_sync(self.file_path)
            if markdown_file:
                # markdown文件非空
                # 将markdown文件处理为文档列表
                docs = convert_markdown_to_langchaindoc(markdown_file)
                docs = self.markdown_process(docs)
                # 构造图片路径
                images_dir = os.path.join(IMAGES_ROOT_PATH, self.file_id)
                # 拷贝图片至指定目录
                self.copy_images(os.path.dirname(markdown_file), images_dir)
            else:
                # markdown文件为空
                # 打印pdf解析异常日志
                insert_logger.warning(
                    f'Error in Powerful PDF parsing, use fast PDF parser instead.')
                # 使用非结构化pdf加载器处理获取文档列表
                loader = UnstructuredPaddlePDFLoader(self.file_path, strategy="fast")
                docs = loader.load()
        elif self.file_path.lower().endswith(".jpg") or self.file_path.lower().endswith(
                ".png") or self.file_path.lower().endswith(".jpeg"):
            # 图片文件，jpg,png,jpeg，后缀名忽略大小写
            # 利用ocr，将图片转换为文本文件，利用TextLoader获取文档列表
            txt_file_path = self.image_ocr_txt(filepath=self.file_path)
            loader = TextLoader(txt_file_path, autodetect_encoding=True)
            docs = loader.load()
        elif self.file_path.lower().endswith(".docx"):
            # word文件
            try:
                # 利用非结构化word文档加载器获取文档列表
                loader = UnstructuredWordDocumentLoader(self.file_path, strategy="fast")
                docs = loader.load()
            except Exception as e:
                # word文件处理异常
                insert_logger.warning('Error in Powerful Word parsing, use docx2txt instead.')
                # 利用docx2txt工具处理，获取文件内容，构造文档对象列表
                text = docx2txt.process(self.file_path)
                docs = [Document(page_content=text)]
        elif self.file_path.lower().endswith(".xlsx"):
            # excel文件
            try:
                # 将excel转为markdown文件，处理markdown文件获取文档列表
                markdown_file = self.excel_to_markdown(self.file_path, os.path.dirname(self.file_path))
                docs = convert_markdown_to_langchaindoc(markdown_file)
                docs = self.markdown_process(docs)
            except Exception as e:
                # excel转markdown异常
                # 打印异常日志
                insert_logger.warning('Error in Powerful Excel parsing, use openpyxl instead.')
                # 初始化文档列表
                docs = []
                # 利用pandas工具，依据excel文件路径获取excel文件
                excel_file = pd.ExcelFile(self.file_path)
                # 获取excel文件的工作簿名称列表
                sheet_names = excel_file.sheet_names
                # 遍历工作波名称列表【枚举，索引+对应索引的元素【也即工作簿名称】】
                for idx, sheet_name in enumerate(sheet_names):
                    # 利用pandas工具读取指定excel文件路径的工作簿
                    xlsx = pd.read_excel(self.file_path, sheet_name=sheet_name, engine='openpyxl')
                    xlsx = xlsx.dropna(how='all', axis=1)  # 只删除全为空的列
                    xlsx = xlsx.dropna(how='all', axis=0)  # 只删除全为空的行
                    # 构造csv文件绝对路径名
                    csv_file_path = self.file_path[:-5] + f'_{idx}.csv'
                    # 将excel工作簿非空行列的数据转为csv文件
                    xlsx.to_csv(csv_file_path, index=False)
                    # 打印csv文件路径
                    insert_logger.info('xlsx2csv: %s', csv_file_path)
                    # 利用csv加载器结合csv文件路径，获取文档列表
                    loader = CSVLoader(csv_file_path, autodetect_encoding=True,
                                       csv_args={"delimiter": ",", "quotechar": '"'})
                    docs.extend(loader.load())
        elif self.file_path.lower().endswith(".pptx"):
            # ppt文件
            # 利用非结构化ppt加载器获取文档列表
            loader = UnstructuredPowerPointLoader(self.file_path, strategy="fast")
            docs = loader.load()
        elif self.file_path.lower().endswith(".eml"):
            # 邮箱文件
            # 利用非结构化邮箱加载器获取文档列表
            loader = UnstructuredEmailLoader(self.file_path, strategy="fast")
            docs = loader.load()
        elif self.file_path.lower().endswith(".csv"):
            # csv文件
            # 利用csv加载器结合csv文件路径，获取文档列表
            loader = CSVLoader(self.file_path, autodetect_encoding=True, csv_args={"delimiter": ",", "quotechar": '"'})
            docs = loader.load()
        else:
            # 其他类型文件，抛出异常
            raise TypeError("文件类型不支持，目前仅支持：[md,txt,pdf,jpg,png,jpeg,docx,xlsx,pptx,eml,csv]")
        # 设置文档元数据并执行合并处理
        self.inject_metadata(docs)

    def inject_metadata(self, docs: List[Document]):
        """
        遍历文档列表，为每个文档设置元数据【用户id、知识库id、文件id、文件名称、文件位置、文件路由、标题列表、是否含有表格、图片路径、页面id、headers、问答字典】
        对于文档内容较小的文档，执行合并处理，最终返回合并后的文档
        """
        # 这里给每个docs片段的metadata里注入file_id
        new_docs = []
        for doc in docs:
            page_content = re.sub(r'\t+', ' ', doc.page_content)  # 将制表符替换为单个空格
            page_content = re.sub(r'\n{3,}', '\n\n', page_content)  # 将三个或更多换行符替换为两个
            page_content = page_content.strip()  # 去除首尾空白字符
            new_doc = Document(page_content=page_content)
            new_doc.metadata["user_id"] = self.user_id
            new_doc.metadata["kb_id"] = self.kb_id
            new_doc.metadata["file_id"] = self.file_id
            new_doc.metadata["file_name"] = self.file_name
            new_doc.metadata["nos_key"] = self.file_location
            new_doc.metadata["file_url"] = self.file_url
            new_doc.metadata["title_lst"] = doc.metadata.get("title_lst", [])
            new_doc.metadata["has_table"] = doc.metadata.get("has_table", False)
            # 从文本中提取图片数量：![figure]（x-figure-x.jpg）
            new_doc.metadata["images"] = re.findall(r'!\[figure]\(\d+-figure-\d+.jpg.*?\)', page_content)
            new_doc.metadata["page_id"] = doc.metadata.get("page_id", 0)
            kb_name = self.mysql_client.get_knowledge_base_name([self.kb_id])[0][2]
            metadata_infos = {"知识库名": kb_name, '文件名': self.file_name}
            new_doc.metadata['headers'] = metadata_infos

            if 'faq_dict' not in doc.metadata:
                new_doc.metadata['faq_dict'] = {}
            else:
                new_doc.metadata['faq_dict'] = doc.metadata['faq_dict']
            new_docs.append(new_doc)
        if new_docs:
            insert_logger.info('langchain analysis content head: %s', new_docs[0].page_content[:100])
        else:
            insert_logger.info('langchain analysis docs is empty!')

        # merge short docs
        insert_logger.info(f"before merge doc lens: {len(new_docs)}")
        # 取默认切片尺寸400和实际切片尺寸一半的最小值
        child_chunk_size = min(DEFAULT_CHILD_CHUNK_SIZE, int(self.chunk_size / 2))
        # 合并的文档列表
        merged_docs = []
        for doc_idx, doc in enumerate(new_docs):
            if not merged_docs:
                # 合并文档列表为空，则将当前文档追加至其中
                merged_docs.append(doc)
            else:
                # 合并文档列表非空，则取其最后一个文档
                last_doc = merged_docs[-1]
                # insert_logger.info(f"doc_idx: {doc_idx}, doc_content: {doc.page_content[:100]}")
                # insert_logger.info(f"last_doc_len: {num_tokens_embed(last_doc.page_content)}, doc_len: {num_tokens_embed(doc.page_content)}")
                if num_tokens_embed(last_doc.page_content) + num_tokens_embed(doc.page_content) <= child_chunk_size or \
                        num_tokens_embed(doc.page_content) < child_chunk_size / 4:
                    # 最后一个文档和当前文档的内容token数量之和不超过child_chunk_size时，或者当前文档的内容token数量少于child_chunk_size的1/4时
                    # 将当前文档的内容按照换行符切分，得到临时内容切片列表
                    tmp_content_slices = doc.page_content.split('\n')
                    # print(last_doc.metadata['title_lst'], tmp_content)
                    # 收集非最后一个文档标题行【有种相对最后一个文档去重的意思】的行内容列表
                    tmp_content_slices_clear = [line for line in tmp_content_slices if clear_string(line) not in
                                                [clear_string(t) for t in last_doc.metadata['title_lst']]]
                    # 利用换行符将上述行内容列表拼接为临时文档内容
                    tmp_content = '\n'.join(tmp_content_slices_clear)
                    # for title in last_doc.metadata['title_lst']:
                    #     tmp_content = tmp_content.replace(title, '')
                    # 将临时文档内容追加到最后一个文档内容中
                    last_doc.page_content += '\n\n' + tmp_content
                    # for title in last_doc.metadata['title_lst']:
                    #     last_doc.page_content = self.remove_substring_after_first(last_doc.page_content, '![figure]')
                    # 将当前文档元数据中的标题列表追加到最后一个文档元数据的标题列表中
                    last_doc.metadata['title_lst'] += doc.metadata.get('title_lst', [])
                    # 将当前文档元数据中的表格标识合并到最后一个文档元数据的表格标识上，逻辑或
                    last_doc.metadata['has_table'] = last_doc.metadata.get('has_table', False) or doc.metadata.get(
                        'has_table', False)
                    # 将当前文档元数据中的图片列表追加到最后一个文档元数据的图片列表中
                    last_doc.metadata['images'] += doc.metadata.get('images', [])
                else:
                    # 最后一个文档和当前文档的内容token数量之和超过child_chunk_size且当前文档的内容token数量不少于child_chunk_size的1/4时
                    # 将当前文档收集到合并文档列表中【相当于没有合并】
                    merged_docs.append(doc)
        # 打印合并后的文档列表长度
        insert_logger.info(f"after merge doc lens: {len(merged_docs)}")
        # 更新切片所得文档列表
        self.docs = merged_docs
